from tkinter import *
from numpy import *
import pandas as pd
import os
import ctypes
import openpyxl
from openpyxl import load_workbook
from tkinter import ttk
import tkinter as tk
from tkinter import filedialog
from tkinter.filedialog import askopenfilename
import xlrd
from tkinter import messagebox
from collections import Counter
import xlsxwriter
from tkinter.filedialog import asksaveasfile
import matplotlib.pyplot as plt
import openpyxl as op
from openpyxl.drawing.image import Image


def save_file():
    
    # opening file explorer to allow user to save file in custom location with custom name
    save_filename = asksaveasfile(initialfile = 'Untitled.xlsx',
    defaultextension=".xlsx",filetypes=[("Excel File","*.xlsx")])

    # saving filename that user created
    save_filename = str(save_filename)
    save_filename = save_filename[25:-29] # data type after saving is not str so after converting it's necessary to modify string to be pure filepath
    save_filename.replace(' ', '-')
    save_filename[26:-31]
    return save_filename
  
  
def data_analyzis(cell_id_list,date_list,line_list,columnnames,defect_by_type):

    overall_defect_count = len(cell_id_list)
    print(overall_defect_count)

    # checking rows and columns length to know 2D array size
    rows= overall_defect_count
    cols = len(columnnames)
    # 2D array for raw data reading
    array_2d = [['' for i in range(cols)] for j in range(rows)]
    i = 0
    for x in cell_id_list:
        array_2d[i][0] = cell_id_list[i]
        array_2d[i][1] = date_list[i]
        array_2d[i][2] = line_list[i]
        array_2d[i][3] = defect_by_type[i]
        i += 1

    # creating date list to read defect count by date
    days_list = set(date_list)
    defect_by_date = ['']*len(days_list)
    
    i = 0
    for x in days_list:
        count = 0
        for day in date_list:
            if day == x:
                count +=1

        defect_by_date[i] = f'{x} : {count}'
        i+=1

    print (defect_by_date) # print for review

    #defect by date and line

    # creating lists for defect per date 
    defect_by_dateline1 = ['']*len(days_list)
    defect_by_dateline2 = ['']*len(days_list)

    # counting defects on lines by day it ocurred
    i = 0
    for i, date in enumerate(days_list):
        count2 = sum(1 for j in range(len(date_list)) if date_list[j] == date and line_list[j] == 2)
        defect_by_dateline2[i] = count2

    for i, date in enumerate(days_list):
        count1 = sum(1 for j in range(len(date_list)) if date_list[j] == date and line_list[j] == 1)
        defect_by_dateline1[i] = count1

    # print to check if it's correct
    print ('line 2: \n', defect_by_dateline2)
    print ('line 1: \n', defect_by_dateline1)

    # filename of newly created excel summary file that thata will be saved to
    save_filename= save_file()
    print(save_filename)


    # set to get non-reapeating date list
    date_set = set(date_list)

    # creating blank 2D array as a data structure to write down to summary excel file
    data_table_for_excel = [['' for i in range(50)] for j in range(len(date_set)+20)]

    # set to get non-reapeating defect list
    defect_by_type_set = set(defect_by_type)

    # print to check if it's correct
    print( '\n',defect_by_type_set,'/n')

 
    # sign headers
    data_table_for_excel[0][0],data_table_for_excel[0][10] = 'line 1','line 2'
    data_table_for_excel[1][0],data_table_for_excel[1][10] = 'date','date' 
    data_table_for_excel[1][1],data_table_for_excel[1][11] = 'defects/day','defects/day' 
    data_table_for_excel[1][3],data_table_for_excel[1][13] = 'defect/type','defect/type'
    data_table_for_excel[1][3],data_table_for_excel[1][13] = 'defect type','defect type'
    
    # date line1 input to 2D array
    i = 2
    for x in date_set:
        data_table_for_excel[i][0] = x
        i += 1

    # defectnumber line1 input to 2D array
    i = 2
    for x in date_set:
        data_table_for_excel[i][1] = defect_by_dateline1[i-2]
        i+= 1

    # line 2 data input to 2D array
    i = 2
    for x in date_set:
        data_table_for_excel[i][10] = x
        i += 1
        
    # defectnumber line 2 input to 2D array
    i = 2
    for x in date_set:
        data_table_for_excel[i][11] = defect_by_dateline2[i-2]
        i+= 1

    # inserting defect type data to 2D array by line
    pos = 2
    for a in defect_by_type_set:
        
        i = 0
        count1 = 0
        count2 = 0
        
        for x in defect_by_type:
            if x == a and array_2d[i][2] == 2:
                count2 +=1              
            elif x == a and array_2d[i][2] == 1:            
                count1 +=1
            i +=1
        
        # inserting defect amount by type on line
        data_table_for_excel[pos][4],data_table_for_excel[pos][14] = count1, count2
        
        # print for correctness revision
        print('\nline1: ', count1, '\nline2', count2)
        pos +=1

    # inserting list of ocurring defect types to 2D array
    i = 2
    for x in defect_by_type_set:
        data_table_for_excel[i][3],data_table_for_excel[i][13] = x,x
        i += 1        
    

    # creating list of defect types by line 2
    defect_by_type_line2 = [[x for x in defect_by_type_set],['' for x in range(len(defect_by_type_set))]]
    
    # inserting defect data by line to list
    i=0
    for x in range(len(defect_by_type_set)):
        defect_by_type_line2[1][i] = data_table_for_excel[i+2][4]
        i += 1
    
    # creating list of defect types by line 1
    defect_by_type_line1 = [[x for x in defect_by_type_set],['' for x in range(len(defect_by_type_set))]]
 
    # inserting defect data by line to list
    i=0
    for x in range(len(defect_by_type_set)):
        defect_by_type_line1[1][i] = data_table_for_excel[i+2][14]
        i += 1

    # print to check correctness
    print(' \n \n 2:', defect_by_type_line2, '\n\n\n\n1:', defect_by_type_line1, '\n\n\n')

    
    # seeking greatest value in defect types lists to find which defect occurs most often on line 1
    max_value_1 = max(defect_by_type_line1[1])
    index_1 = defect_by_type_line1[1].index(max_value_1)
    
    # printing main defect type for line 1
    data_table_for_excel[len(date_set)+3][10], data_table_for_excel[len(date_set)+4][23] = 'Main defect type line 1:', f'{defect_by_type_line1[0][index_1]}'

    # seeking greatest value in defect types lists to find which defect occurs most often on line 2
    max_value_2 = max(defect_by_type_line2[1])
    index_2 = defect_by_type_line2[1].index(max_value_2)

    # printing main defect type for line 2
    data_table_for_excel[len(date_set)+3][23], data_table_for_excel[len(date_set)+4][10] = 'Main defect type line 2:', f'{defect_by_type_line2[0][index_2]}'
 

    df = pd.DataFrame(data_table_for_excel)
    df.to_excel(save_filename, index = False, header= False)

 
    # creating pie chart for line 1 for visualization percentage amount of defects on line
    data = {'Defect Type': defect_by_type_line1[0],
        'Percentage amount of defects types - line 1': defect_by_type_line1[1]}

    plt.figure(figsize=(8, 8))

    plt.pie(data['Percentage amount of defects types - line 1'], labels=data['Defect Type'], autopct='%1.1f%%')

    plt.title('Percentage amount of defects types - line 2')

    plt.savefig(f'circle_chart1.png' ) # saving chart to .png file in script localization

    plt.close()

    # declaration of workbook that chart is going to be write to
    wb = load_workbook(save_filename)
    sheet = wb.active

    # it is important to declare last row to keep charts below data
    last_row = sheet.max_row

    img1 = Image(f'circle_chart1.png' )
    img1.width = 600  
    img1.height = 600  

    # Placing the chart below existing data 
    chart_location1 = f'N{last_row-1}'  
    sheet.add_image(img1, chart_location1)

    # saving workbook changes
    wb.save(save_filename)


    # creating pie chart for line 1 for visualization percentage amount of defects on line 2
    data = {'Defect Type': defect_by_type_line2[0],
        'Percentage amount of defects types - line 2': defect_by_type_line2[1]}

    plt.figure(figsize=(8, 8))

    plt.pie(data['Percentage amount of defects types - line 2'], labels=data['Defect Type'], autopct='%1.1f%%')

    plt.title('Percentage amount of defects types - line 1')

    plt.savefig(f'circle_chart2.png') # saving chart to .png file in script localization

    plt.close()

    # declaration of workbook that chart is going to be write to
    wb = load_workbook(save_filename)
    sheet = wb.active


    img2 = Image(f'circle_chart2.png')
    img2.width = 600  
    img2.height = 600  
 
    # Placing the chart below existing data 
    chart_location2 = f'A{last_row-1}'                           
    sheet.add_image(img2, chart_location2)

    # saving workbook changes
    wb.save(save_filename)

    # saving changes to summary excel file
    f = os.system(f'start "excel" "{save_filename}" ')
    
    
def data_structure(columnnames, doc_name,window):
    
    ds =  pd.read_excel(doc_name, usecols='A:D', index_col=None)
    ds.fillna(0, inplace = True)

    defect_by_type = ds[ds.columns[3]].values.tolist()

    cell_id_list = ds[ds.columns[0]].values.tolist()

    date_list = ds[ds.columns[1]].values.tolist()

    line_list = ds[ds.columns[2]].values.tolist()
    
    data_analyzis(cell_id_list,date_list,line_list,columnnames,defect_by_type)
    ######  Data structure ready. important to add proper reading for analizys      #######
 
 
def button_onclick(doc_name,window):
    print(doc_name)

    df = pd.read_excel(doc_name, na_filter=True)
    columnnames = df.columns.values.tolist()

    print (columnnames)

    data_structure(columnnames, doc_name,window)


def button_generator(doc_name,window):

    # adding a caption of which file is opened for analysis
    openlabel = Label(window, text=f"\nYou are opening: {doc_name}. \n \nTo analyze the file click on button with it's filename below \n", width=100)
    openlabel.grid(row=4, column = 0, columnspan= 100)       

    emptylabel = Label(window, text='\n', width=100)
    emptylabel.grid(row=6, column = 0, columnspan= 50)       

    button = tk.Button(window,width='50',bd= 4,bg='#FFF6DA', justify=CENTER, text= doc_name, command = lambda : button_onclick(doc_name,window))
    button.grid(row=5, column = 0)


def browseFiles(window):

    filename = filedialog.askopenfilename(initialdir = "/",
                                          title = "Select a File",
                                          filetypes = (('Excel files',
                                                         '*.xlsx*'),
                                                        ("Text files",
                                                        "*.txt*"),
                                                       ("all files",
                                                        "*.*")))
    doc_name =filename
    button_generator(doc_name,window)
    return filename


def window():
    # declaring tkinter window
    window = tk.Tk()
    window.title = ('excel transformation')
    window.geometry = ('1200x680')

    # in this project tkinter frame method is used because it's easier to implement in this case
    frame = ttk.Frame(window, padding = 10)
    frame.grid()

    my_str = tk.StringVar()
    l1 = tk.Label(window, textvariable=my_str, width = 50)
    l1.grid(row=0, column = 0, columnspan= 50)

    header = Label(window, text='Excel data analyze \n',font=('ARIAL', 24), width=25)
    header.grid(row=1, column = 0, columnspan= 100)

    l2 = Label(window, text='To browse desired file please click on \"browse\" button \n',font=('ARIAL', 16), width=50)
    l2.grid(row=2, column = 0, columnspan= 50)

    button_exit = tk.Button(window, text='exit', justify=CENTER,width='20',bg='#C0C0C0', command = lambda : exit())
    button_exit.grid(column=0, row=15)

    button_browse = Button(window, text = 'browse',width='100', command = lambda : browseFiles(window))
    button_browse.grid( column=0, row= 3)

    window.mainloop()

# run window
window()